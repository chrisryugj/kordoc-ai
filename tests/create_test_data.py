"""테스트용 샘플 데이터 생성 스크립트.

실행: python tests/create_test_data.py
결과: tests/test_data/ 아래에 MVP별 테스트 파일 생성
"""

import os
import sys
from pathlib import Path

# 프로젝트 루트를 경로에 추가
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

TEST_DATA = ROOT / "tests" / "test_data"


def create_mvp1_test_pdf():
    """MVP 1 테스트용 간단한 PDF 생성 (reportlab 사용)."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        print("  [SKIP] reportlab 미설치 - MVP1 테스트 PDF 생략")
        return False

    pdf_dir = TEST_DATA / "mvp1"
    pdf_path = pdf_dir / "test_document.pdf"

    # 한글 폰트
    font_name = "MalgunGothic"
    try:
        pdfmetrics.registerFont(TTFont(font_name, r"C:\Windows\Fonts\malgun.ttf"))
    except Exception:
        font_name = "Helvetica"

    c = canvas.Canvas(str(pdf_path))

    # 페이지 1: 텍스트
    c.setFont(font_name, 16)
    c.drawString(50, 750, "사전기획 적정성 검토 보고서")
    c.setFont(font_name, 12)
    c.drawString(50, 720, "작성일: 2026년 3월 20일")
    c.drawString(50, 700, "작성자: 교육시설지원처 사전기획팀")
    c.drawString(50, 660, "1. 사업 개요")
    c.drawString(70, 640, "본 사업은 노후 교육시설의 공간 재구조화를 통해")
    c.drawString(70, 620, "미래형 교육환경을 조성하는 것을 목적으로 합니다.")
    c.drawString(50, 580, "2. 추진 경과")
    c.drawString(70, 560, "- 2025.03: 사업 대상 학교 선정 (15개교)")
    c.drawString(70, 540, "- 2025.06: 기본계획 수립 완료")
    c.drawString(70, 520, "- 2025.09: 설계 용역 착수")
    c.drawString(50, 480, "3. 예산 현황")
    c.drawString(70, 460, "총 사업비: 150억원")
    c.drawString(70, 440, "집행액: 45억원 (30%)")
    c.showPage()

    # 페이지 2: 표 형태
    c.setFont(font_name, 14)
    c.drawString(50, 750, "학교별 사업 현황")
    c.setFont(font_name, 10)
    y = 720
    headers = ["학교명", "사업유형", "면적(m2)", "예산(억)"]
    data = [
        ["광진초등학교", "전면 리모델링", "2,500", "35"],
        ["자양중학교", "부분 개선", "1,200", "18"],
        ["건국고등학교", "증축", "800", "22"],
        ["광장초등학교", "그린스마트", "3,100", "42"],
        ["중곡중학교", "복합화", "1,800", "33"],
    ]
    # 헤더
    for i, h in enumerate(headers):
        c.drawString(50 + i * 130, y, h)
    y -= 20
    c.line(50, y + 10, 550, y + 10)
    # 데이터
    for row in data:
        for i, cell in enumerate(row):
            c.drawString(50 + i * 130, y, cell)
        y -= 18
    c.showPage()

    # 페이지 3: 짧은 내용
    c.setFont(font_name, 12)
    c.drawString(50, 750, "4. 향후 계획")
    c.drawString(70, 720, "2026년 상반기 내 설계 완료 및 착공 예정")
    c.drawString(70, 700, "공간 재구조화 가이드라인에 따른 품질 관리 강화")
    c.showPage()

    c.save()
    print(f"  [OK] MVP1 테스트 PDF: {pdf_path} (3페이지)")
    return True


def create_mvp2_test_data():
    """MVP 2 테스트용 PDF + Excel 생성."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import openpyxl
    except ImportError:
        print("  [SKIP] reportlab/openpyxl 미설치 - MVP2 테스트 데이터 생략")
        return False

    mvp2_dir = TEST_DATA / "mvp2"
    source_dir = mvp2_dir / "source"

    font_name = "MalgunGothic"
    try:
        pdfmetrics.registerFont(TTFont(font_name, r"C:\Windows\Fonts\malgun.ttf"))
    except Exception:
        font_name = "Helvetica"

    # 원본 PDF 2개 생성 (각 5페이지)
    for idx, name in enumerate(["1_사업계획서.pdf", "2_보고서.pdf"], 1):
        pdf_path = source_dir / name
        c = canvas.Canvas(str(pdf_path))
        for pg in range(1, 6):
            c.setFont(font_name, 14)
            c.drawString(50, 750, f"{name} - {pg}페이지")
            c.setFont(font_name, 11)
            c.drawString(50, 720, f"문서 {idx}의 {pg}번째 페이지 내용입니다.")
            c.drawString(50, 700, f"테스트 데이터 - 테마 추출 검증용")
            if pg == 1:
                c.drawString(50, 660, "학급수: 24학급, 학생수: 650명")
            elif pg == 2:
                c.drawString(50, 660, "대지면적: 12,500m2, 건물면적: 8,200m2")
            elif pg == 3:
                c.drawString(50, 660, "사업유형: 그린스마트 미래학교")
            elif pg == 4:
                c.drawString(50, 660, "사업기간: 2025.03 ~ 2027.02 (24개월)")
            elif pg == 5:
                c.drawString(50, 660, "총 사업비: 350억원 (국비 200억, 시비 150억)")
            c.showPage()
        c.save()
        print(f"  [OK] MVP2 원본 PDF: {pdf_path} (5페이지)")

    # 엑셀 인덱스 생성
    xlsx_path = mvp2_dir / "pages.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더행 (1행)
    ws.cell(row=1, column=1, value="테마")
    ws.cell(row=1, column=2, value="설명")
    ws.cell(row=1, column=3, value="문서1 페이지")
    ws.cell(row=1, column=4, value="문서2 페이지")

    # 데이터 (settings.yaml의 topic_folders 행번호와 일치)
    topics = {
        2: ("a_학급수 및 학생수", "학급/학생 현황", "1", "1"),
        3: ("b_각종 면적", "면적 정보", "2", "2"),
        4: ("c_사업 유형", "사업 유형 분류", "3", "3"),
        5: ("d_사업 기간", "사업 기간 정보", "4", "4"),
        6: ("e_설계발주방식", "설계 발주", "", ""),
        7: ("f_사업비 내역", "예산 내역", "5", "5"),
        8: ("g_스페이스 프로그램", "공간 프로그램", "", ""),
        9: ("h_설문조사 결과", "설문 결과", "", ""),
    }
    for row_num, (folder, desc, pg_c, pg_d) in topics.items():
        ws.cell(row=row_num, column=1, value=folder)
        ws.cell(row=row_num, column=2, value=desc)
        ws.cell(row=row_num, column=3, value=pg_c)
        ws.cell(row=row_num, column=4, value=pg_d)

    wb.save(xlsx_path)
    print(f"  [OK] MVP2 엑셀 인덱스: {xlsx_path}")
    return True


def create_mvp3_test_data():
    """MVP 3 테스트용 교육과정 텍스트 파일 생성."""
    input_dir = TEST_DATA / "mvp3" / "input_data"

    schools = {
        "광진초등학교.txt": """
광진초등학교 교육과정 운영 계획 (2025학년도)

1. 학교 현황
- 학급수: 24학급 (1~6학년 각 4학급)
- 학생수: 650명
- 교직원: 48명

2. 특색교육
가. 생태환경교육 "푸른 광진"
- 학교 텃밭 운영 (전 학년 참여)
- 탄소중립 실천 프로젝트
- 지역 하천(중랑천) 생태 탐사 프로그램

나. AI 디지털 리터러시 교육
- 3~6학년 코딩 교육 주 1회
- AI 윤리 교육 연 4회
- 디지털 시민성 프로그램

3. 교육과정 운영 방식
- 블록타임제 운영 (80분 블록)
- 학년군별 교육과정 재구성
- 프로젝트 기반 학습(PBL) 도입
""",
        "자양중학교.txt": """
자양중학교 교육과정 운영 계획 (2025학년도)

1. 학교 현황
- 학급수: 18학급 (1~3학년 각 6학급)
- 학생수: 520명
- 교직원: 42명

2. 특색교육
가. 국제이해교육 "세계 시민 자양"
- 자매결연 해외 학교 교류 (일본, 대만)
- 다문화 이해 주간 운영 (연 2회)
- 영어 몰입 캠프 (방학 중)

나. 메이커 교육
- 메이커 스페이스 운영 (3D 프린터, 레이저 커터)
- 발명 동아리 "자양 이노베이터"
- 지역 연계 창작 페스티벌 참가

3. 교육과정 운영 방식
- 자유학기제 (1학년 2학기)
- 선택 교과제 (2~3학년, 12과목)
- 교과 융합 수업 (STEAM)
""",
        "건국고등학교.txt": """
건국고등학교 교육과정 운영 계획 (2025학년도)

1. 학교 현황
- 학급수: 30학급 (1~3학년 각 10학급)
- 학생수: 890명
- 교직원: 65명

2. 특색교육
가. 진로집중교육 "건국 패스파인더"
- 학과 탐색 프로그램 (대학 연계)
- 1인 1진로 포트폴리오
- 직업 체험 프로그램 (학기당 2회)

나. 독서·토론 교육
- 아침 독서 시간 (주 3회, 20분)
- 학년별 필독서 선정 및 독서 토론대회
- 논술 클리닉 (3학년 대상)

3. 교육과정 운영 방식
- 고교학점제 선도학교
- 공동교육과정 운영 (인근 3개교)
- AI 보조 학습 시스템 도입
""",
    }

    for filename, content in schools.items():
        filepath = input_dir / filename
        filepath.write_text(content.strip(), encoding="utf-8")
        print(f"  [OK] MVP3 텍스트: {filepath}")

    return True


def create_test_config():
    """테스트용 settings.yaml 오버라이드 생성."""
    config_path = TEST_DATA / "test_settings.yaml"
    # 상대경로 사용 — CI/로컬 모두 동작 (테스트 코드에서 절대경로로 변환)
    content = """# 테스트용 설정 (tests/test_data 기준 상대경로)
mvp1:
  target_dir: "mvp1"
  model_name: "gemini-3.1-flash"
  temperature: 0.0
  max_workers: 2
  timeout_seconds: 60
  chunk_size: 10
  dpi: 100
  max_retries: 2
  base_delay: 5

mvp2:
  base_dir: "mvp2"
  source_subdir: "source"
  index_file: "pages.xlsx"
  source_files:
    C: "1_사업계획서.pdf"
    D: "2_보고서.pdf"
  topic_folders:
    2: "a_학급수 및 학생수"
    3: "b_각종 면적"
    4: "c_사업 유형"
    5: "d_사업 기간"
    6: "e_설계발주방식"
    7: "f_사업비 내역"
    8: "g_스페이스 프로그램"
    9: "h_설문조사 결과"
  label:
    font_size: 10
    x: 20
    y: 20
    color: [0, 0, 0]

mvp3:
  base_dir: "mvp3"
  input_subdir: "input_data"
  processed_subdir: "input_processed"
  output_subdir: "output"
  model_name: "gemini-3.1-flash"
  step1_temperature: 0.0
  step2_temperature: 0.7
  timeout_seconds: 120
  wait_time: 5
  num_variants: 1
"""
    config_path.write_text(content, encoding="utf-8")
    print(f"  [OK] 테스트 설정: {config_path}")
    return True


if __name__ == "__main__":
    print("=" * 50)
    print("  테스트 데이터 생성")
    print("=" * 50)
    print()
    create_mvp1_test_pdf()
    create_mvp2_test_data()
    create_mvp3_test_data()
    create_test_config()
    print()
    print("완료!")
